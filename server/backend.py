"""
Agent Backend
Runs on the host machine and interacts with the test_clawd container via docker exec.
Supports: Z-Image generation / Auto-Test inference / Auto-Quant quantization / Auto-Eval evaluation
No SSH or asyncssh required.

Start:
    pip install fastapi uvicorn
    cd /home/kaokao/wenjiao_latest/open-claw/script
    uvicorn back_end:app --host 0.0.0.0 --port 8000
"""

import asyncio
import json
import logging
import os
import shlex
import time
import uuid

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

DOCKER_CONTAINER   = "test_clawd"
WORKDIR            = "/kaokao/openclaw-triton-gen/examples/auto_run"    # inside docker
WORKDIR_AUTOQUANT  = "/kaokao/openclaw-triton-gen/examples/auto_run"    # inside docker
WORKDIR_AUTOEVAL   = "/kaokao/openclaw-triton-gen/examples/auto_run"    # inside docker
AUTORUN_SKILL      = "/root/.openclaw/workspace/skills/auto_run/SKILL.md"   # inside docker
AUTOQUANT_SKILL    = "/root/.openclaw/workspace/skills/auto_quant/SKILL.md" # inside docker
AUTOEVAL_SKILL     = "/root/.openclaw/workspace/skills/auto_eval/SKILL.md"  # inside docker
EXCEL_XLX_SKILL    = "/wenjiao/openclaw-triton-gen/skills/excel-xlsx/SKILL.md" # inside docker
ZIMAGE_LOG         = "/storage/lkk/inference/zimage/logs/zimage.log"   # inside docker
SESSION_DIR        = "/root/.openclaw/agents/main/sessions"             # inside docker

PROXY_EXPORTS = (
    "export https_proxy=http://child-igk.intel.com:912 && "
    "export http_proxy=http://child-igk.intel.com:912"
)

_HERE   = os.path.dirname(os.path.abspath(__file__))

app = FastAPI()
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

try:
    app.mount("/static", StaticFiles(directory=_HERE), name="static")
except Exception:
    pass


task_queues: dict[str, asyncio.Queue] = {}


def _evt(task_id, type_, message, stage="zimage", level="info", meta=None):
    return {
        "task_id":      task_id,
        "type":         type_,
        "message":      message,
        "stage":        stage,
        "level":        level,
        "timestamp_ms": int(time.time() * 1000),
        "meta":         meta or {},
    }


class MachineConfig(BaseModel):
    # SSH / remote machine (optional, for future SSH-based backends)
    host:        str = "h3c.sh.intel.com"
    user:        str = "kaokao"
    password:    str = ""
    # Docker & paths
    container:   str = DOCKER_CONTAINER
    repo_root:   str = "/kaokao/openclaw-triton-gen"
    workdir:     str = WORKDIR
    output_root: str = "/storage/lkk/inference"
    session_dir: str = SESSION_DIR
    minimax_key: str = ""


class ZImageRequest(BaseModel):
    prompt:              str = "a cup of coffee on the table"
    output:              str = "/storage/lkk/inference/zimage/zimage_output.png"
    num_inference_steps: int = 9
    seed:                int = 42
    device:              str = "xpu"
    session_key:         str = "zimage_task"
    machine:             MachineConfig = MachineConfig()


class ExcelXLXRequest(BaseModel):
    file_path:   str = ""   # Excel file path inside container (optional)
    instructions: str = ""  # what to do
    skill_path:  str = EXCEL_XLX_SKILL
    session_key: str = "excel_xlx_task"
    timeout:     int = 300
    machine:     MachineConfig = MachineConfig()


class AutoTestRequest(BaseModel):
    model_id:     str = "Qwen/Qwen3-0.6B"
    device:       str = "cuda"
    device_index: str = "0"
    output_dir:   str = "/storage/lkk/inference"
    skill_path:   str = AUTORUN_SKILL
    session_key:  str = "autotest_task"
    timeout:      int = 7200
    machine:      MachineConfig = MachineConfig()


class AutoQuantRequest(BaseModel):
    model_id:      str = "Qwen/Qwen3-0.6B"
    scheme:        str = "W4A16"
    method:        str = "RTN"
    export_format: str = "auto_round:auto_gptq"
    device:        str = "xpu"
    device_index:  str = "0"
    output_dir:    str = "/kaokao/quantized"
    workspace_dir: str = ""   # auto_run output dir for this model (contains model_info.json)
    skill_path:    str = AUTOQUANT_SKILL
    session_key:   str = "autoquant_task"
    timeout:       int = 7200
    machine:       MachineConfig = MachineConfig()


class AutoEvalRequest(BaseModel):
    model_path:              str   = "/kaokao/quantized/Qwen_Qwen3-0.6B-W4A16"
    tasks:                   str   = "piqa"
    output_dir:              str   = "/kaokao/lm_eval_results/Qwen_Qwen3-0.6B"
    device:                  str   = "cuda"
    device_index:            str   = "0"
    batch_size:              int   = 8
    max_model_len:           int   = 8192
    gpu_memory_utilization:  float = 0.8
    skill_path:              str   = AUTOEVAL_SKILL
    session_key:             str   = "autoeval_task"
    timeout:                 int   = 7200
    machine:                 MachineConfig = MachineConfig()


class PipelineRequest(BaseModel):
    model_id:      str  = "Qwen/Qwen3-0.6B"
    device:        str  = "cuda"
    device_index:  str  = "0"
    timeout:       int  = 7200
    session_key:   str  = "pipeline"

    run_autotest:  bool = False
    run_autoquant: bool = True
    run_autoeval:  bool = True

    # Auto-Test
    test_output_dir:  str = "/storage/lkk/inference"
    autorun_skill:    str = AUTORUN_SKILL

    # Auto-Quant (also used to derive eval model path)
    scheme:           str = "W4A16"
    method:           str = "RTN"
    export_format:    str = "auto_round:auto_gptq"
    quant_output_dir: str = "/kaokao/quantized"
    autoquant_skill:  str = AUTOQUANT_SKILL

    # Auto-Eval
    tasks:                   str   = "piqa"
    eval_output_dir:         str   = "/kaokao/lm_eval_results"
    batch_size:              int   = 8
    max_model_len:           int   = 8192
    gpu_memory_utilization:  float = 0.8
    skill_path:              str   = AUTOEVAL_SKILL
    machine:                 MachineConfig = MachineConfig()


@app.post("/api/run-zimage")
async def run_zimage(req: ZImageRequest):
    task_id = str(uuid.uuid4())
    task_queues[task_id] = asyncio.Queue()
    asyncio.create_task(run_zimage_task(task_id, req))
    return {"task_id": task_id}


@app.post("/api/run-autotest")
async def run_autotest(req: AutoTestRequest):
    task_id = str(uuid.uuid4())
    task_queues[task_id] = asyncio.Queue()
    asyncio.create_task(run_autotest_task(task_id, req))
    return {"task_id": task_id}


@app.post("/api/run-autoquant")
async def run_autoquant(req: AutoQuantRequest):
    task_id = str(uuid.uuid4())
    task_queues[task_id] = asyncio.Queue()
    asyncio.create_task(run_autoquant_task(task_id, req))
    return {"task_id": task_id}


@app.post("/api/run-autoeval")
async def run_autoeval(req: AutoEvalRequest):
    task_id = str(uuid.uuid4())
    task_queues[task_id] = asyncio.Queue()
    asyncio.create_task(run_autoeval_task(task_id, req))
    return {"task_id": task_id}


@app.post("/api/run-pipeline")
async def run_pipeline(req: PipelineRequest):
    task_id = str(uuid.uuid4())
    task_queues[task_id] = asyncio.Queue()
    asyncio.create_task(run_pipeline_task(task_id, req))
    return {"task_id": task_id}


async def _docker_exec(cmd_inside: str, container: str = DOCKER_CONTAINER) -> "asyncio.subprocess.Process":
    """Spawn docker exec and return the process (stdout piped)."""
    return await asyncio.create_subprocess_exec(
        "docker", "exec", container, "bash", "-c", cmd_inside,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.STDOUT,
    )


async def _wait_file_in_docker(path: str, stop: asyncio.Event, timeout: float = 120, container: str = DOCKER_CONTAINER) -> bool:
    deadline = time.time() + timeout
    while time.time() < deadline:
        if stop.is_set():
            return False
        p = await asyncio.create_subprocess_exec(
            "docker", "exec", container, "test", "-f", path,
            stdout=asyncio.subprocess.DEVNULL,
            stderr=asyncio.subprocess.DEVNULL,
        )
        await p.wait()
        if p.returncode == 0:
            return True
        await asyncio.sleep(0.5)
    return False


async def _tail_file(queue, task_id, path_in_docker, evt_type, stop, container: str = DOCKER_CONTAINER):
    """tail -n 0 -f a file inside docker, emit events."""
    if not await _wait_file_in_docker(path_in_docker, stop, container=container):
        return
    proc = await asyncio.create_subprocess_exec(
        "docker", "exec", container,
        "tail", "-n", "0", "-f", path_in_docker,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.DEVNULL,
    )
    try:
        async for raw in proc.stdout:
            if stop.is_set():
                break
            line = raw.decode(errors="replace").rstrip()
            if line:
                await queue.put(_evt(task_id, evt_type, line, stage=evt_type))
    finally:
        try:
            proc.kill()
        except Exception:
            pass


def _parse_session_line(line: str) -> list[tuple[str, str]]:
    """Parse one JSONL line from an openclaw session file.
    Returns list of (sub_type, text) tuples to emit."""
    results = []
    try:
        obj = json.loads(line)
        msg = obj.get("message")
        if not msg:
            return results
        role = msg.get("role", "")
        if role == "user":
            return results
        content = msg.get("content") or []
        if isinstance(content, str):
            content = [{"type": "text", "text": content}]
        for c in content:
            t = c.get("type", "")
            if t == "thinking":
                results.append(("thinking", c.get("thinking", "")[:800]))
            elif t == "text":
                results.append(("text", c.get("text", "")[:8000]))
            elif t == "toolCall":
                args = c.get("arguments", {})
                name = c.get("name", "?")
                text = f"{name}: {args.get('command', str(args))[:500]}"
                results.append(("tool_call", text))
            elif t == "toolResult":
                r = c.get("content", "")
                if isinstance(r, list):
                    r = "".join(x.get("text", "") for x in r)
                results.append(("tool_result", str(r)[:3000]))
        # toolResult role messages have their content as direct text items
        if role == "toolResult" and not results:
            for c in content:
                if c.get("type") == "text":
                    results.append(("tool_result", c.get("text", "")[:3000]))
    except Exception:
        pass
    return results


async def _watch_and_tail_session(queue, task_id, session_dir: str, stop, container: str = DOCKER_CONTAINER):
    """
    Dynamically find the session JSONL that openclaw is actually writing to,
    then stream its new content as transcript events.

    openclaw names session files by internal UUID (not by --session-id), and
    when --agent main is used it continues the existing agent:main:main session.
    We snapshot sizes before the run, then detect whichever file grows/appears.
    """
    # --- Step 1: snapshot all .jsonl files and their sizes ---
    snap_cmd = (
        f"find {shlex.quote(session_dir)} -maxdepth 1 -name '*.jsonl' "
        f"-exec stat -c '%n %s' {{}} + 2>/dev/null || true"
    )
    snap_proc = await asyncio.create_subprocess_exec(
        "docker", "exec", container, "bash", "-c", snap_cmd,
        stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.DEVNULL,
    )
    out, _ = await snap_proc.communicate()
    before: dict[str, int] = {}
    for ln in out.decode().splitlines():
        parts = ln.rsplit(None, 1)
        if len(parts) == 2:
            before[parts[0]] = int(parts[1])

    # --- Step 2: poll until a file appears or grows ---
    target_file: str | None = None
    start_offset: int = 0
    deadline = time.time() + 60
    while time.time() < deadline and not stop.is_set():
        check_proc = await asyncio.create_subprocess_exec(
            "docker", "exec", container, "bash", "-c", snap_cmd,
            stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.DEVNULL,
        )
        curr_out, _ = await check_proc.communicate()
        best_new: str | None = None   # newly created file
        best_grown: str | None = None  # existing file that grew
        best_grown_offset: int = 0
        for ln in curr_out.decode().splitlines():
            parts = ln.rsplit(None, 1)
            if len(parts) != 2:
                continue
            fname, sz = parts[0], int(parts[1])
            prev = before.get(fname, -1)
            if sz > 0:
                if prev == -1:          # brand-new file
                    best_new = fname
                    break
                elif sz > prev:         # existing file grew
                    if best_grown is None:
                        best_grown = fname
                        best_grown_offset = prev
        if best_new is not None:
            target_file = best_new
            start_offset = 0
            break
        if best_grown is not None:
            target_file = best_grown
            start_offset = best_grown_offset
            break
        await asyncio.sleep(0.5)

    if not target_file or stop.is_set():
        return

    # --- Step 3: tail from start_offset ---
    # tail -c +N reads from byte N (1-indexed); +1 = from beginning, +(size+1) = after existing content
    tail_cmd = f"tail -c +{start_offset + 1} -f {shlex.quote(target_file)}"
    proc = await asyncio.create_subprocess_exec(
        "docker", "exec", container, "bash", "-c", tail_cmd,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.DEVNULL,
    )
    buf = ""
    try:
        async for raw in proc.stdout:
            if stop.is_set():
                break
            buf += raw.decode(errors="replace")
            while "\n" in buf:
                line, buf = buf.split("\n", 1)
                line = line.strip()
                if not line:
                    continue
                for sub_type, text in _parse_session_line(line):
                    await queue.put(_evt(task_id, "transcript", text,
                                        meta={"sub_type": sub_type}))
    finally:
        try:
            proc.kill()
        except Exception:
            pass


async def _tail_session_jsonl(queue, task_id, session_file, stop, container: str = DOCKER_CONTAINER):
    """Legacy: tail a known session file path. Use _watch_and_tail_session when file name is unknown."""
    if not await _wait_file_in_docker(session_file, stop, container=container):
        return
    proc = await asyncio.create_subprocess_exec(
        "docker", "exec", container,
        "tail", "-n", "0", "-f", session_file,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.DEVNULL,
    )
    buf = ""
    try:
        async for raw in proc.stdout:
            if stop.is_set():
                break
            buf += raw.decode(errors="replace")
            while "\n" in buf:
                line, buf = buf.split("\n", 1)
                line = line.strip()
                if not line:
                    continue
                for sub_type, text in _parse_session_line(line):
                    await queue.put(_evt(task_id, "transcript", text,
                                        meta={"sub_type": sub_type}))
    finally:
        try:
            proc.kill()
        except Exception:
            pass


async def run_zimage_task(task_id: str, req: ZImageRequest):
    queue = task_queues[task_id]

    async def emit(type_, msg, stage="zimage", level="info", meta=None):
        await queue.put(_evt(task_id, type_, msg, stage, level, meta))

    try:
        await emit("status", f"Starting Z-Image agent: {req.prompt[:60]!r}")
        container = req.machine.container

        # Clear session file before run (truncate to 0)
        session_file = f"{req.machine.session_dir}/{req.session_key}.jsonl"
        clear_proc = await asyncio.create_subprocess_exec(
            "docker", "exec", container,
            "bash", "-c", f"truncate -s 0 {shlex.quote(session_file)} 2>/dev/null || true",
            stdout=asyncio.subprocess.DEVNULL, stderr=asyncio.subprocess.DEVNULL,
        )
        await clear_proc.wait()

        env_exports = " && ".join([
            "export https_proxy=http://child-igk.intel.com:912",
            "export http_proxy=http://child-igk.intel.com:912",
            "export HF_HOME=/storage/lkk/cache",
            "export PYTHONUNBUFFERED=1",
            f"export ZIMAGE_PROMPT={shlex.quote(req.prompt)}",
            f"export ZIMAGE_OUTPUT={shlex.quote(req.output)}",
            f"export ZIMAGE_STEPS={req.num_inference_steps}",
            f"export ZIMAGE_DEVICE={shlex.quote(req.device)}",
            f"export ZIMAGE_SESSION={shlex.quote(req.session_key)}",
        ])
        inner = f"cd {req.machine.workdir} && {env_exports} && python3 test_zimage_agent.py"

        stop_event = asyncio.Event()

        session_task = asyncio.create_task(
            _tail_session_jsonl(queue, task_id, session_file, stop_event, container=container)
        )
        log_task = asyncio.create_task(
            _tail_file(queue, task_id, ZIMAGE_LOG, "zimage_log", stop_event, container=container)
        )

        proc = await asyncio.create_subprocess_exec(
            "docker", "exec", container, "bash", "-c", inner,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.STDOUT,
        )

        async def _drain():
            async for raw in proc.stdout:
                line = raw.decode(errors="replace").rstrip()
                if line:
                    await emit("log", line)

        await asyncio.gather(_drain(), proc.wait())
        rc = proc.returncode

        await asyncio.sleep(5)
        stop_event.set()
        for t in (session_task, log_task):
            t.cancel()
            try:
                await t
            except asyncio.CancelledError:
                pass

        if rc == 0:
            await emit("done", "Z-Image finished successfully", meta={"output": req.output})
        else:
            await emit("error", f"Z-Image agent exited with code {rc}",
                       level="error", meta={"exit_code": rc})

    except Exception as e:
        logger.exception("run_zimage_task failed")
        await queue.put(_evt(task_id, "error", str(e), level="error"))
    finally:
        await queue.put(None)


# ─────────────────────────── Auto-Test task ───────────────────────────────

def _build_autotest_prompt(req: "AutoTestRequest") -> str:
    device = req.device.lower()
    output_dir = f"{req.output_dir}/{req.model_id.replace('/', '_')}"

    if device == "xpu":
        device_block = (
            f"STRICT XPU-ONLY: set os.environ['ZE_AFFINITY_MASK']='{req.device_index}' as the FIRST line, "
            f"then import torch. Load model with device_map='xpu:{req.device_index}' and torch_dtype=torch.bfloat16. "
            f"Do NOT use CUDA APIs, CUDA_VISIBLE_DEVICES, or CPU fallback."
        )
    elif device == "cuda":
        device_block = (
            f"Use CUDA device cuda:{req.device_index}. "
            f"Set CUDA_VISIBLE_DEVICES={req.device_index} before importing torch."
        )
    else:
        device_block = "Use CPU device."

    return (
        f"You are an expert in LLM inference. Your task is to run a quick smoke test on a large language model.\n"
        f"You MUST follow the skill instructions in: {req.skill_path}\n\n"
        f"Model: {req.model_id}\n"
        f"Target device: {device}:{req.device_index}\n"
        f"Output directory: {output_dir}\n\n"
        f"Device requirement: {device_block}\n\n"
        f"System environment (concrete values for this machine):\n"
        f"- Shell: always use bash; use 'bash -c' explicitly in shell commands.\n"
        f"- Base Python: /kaokao/miniforge3/bin/python3  (has torch 2.9.0+xpu pre-installed)\n"
        f"- HuggingFace cache dir: HF_HOME=/storage/lkk/cache\n"
        f"- Network proxy: https_proxy=http://child-igk.intel.com:912  http_proxy=http://child-igk.intel.com:912\n"
    )


async def run_autotest_task(task_id: str, req: "AutoTestRequest"):
    queue = task_queues[task_id]

    async def emit(type_, msg, stage="autotest", level="info", meta=None):
        await queue.put(_evt(task_id, type_, msg, stage, level, meta))

    try:
        await emit("status", f"Starting Auto-Test: {req.model_id} on {req.device}:{req.device_index}")
        container = req.machine.container

        session_file = f"{req.machine.session_dir}/{req.session_key}.jsonl"
        clear_proc = await asyncio.create_subprocess_exec(
            "docker", "exec", container,
            "bash", "-c", f"truncate -s 0 {shlex.quote(session_file)} 2>/dev/null || true",
            stdout=asyncio.subprocess.DEVNULL, stderr=asyncio.subprocess.DEVNULL,
        )
        await clear_proc.wait()

        task_prompt = _build_autotest_prompt(req)
        inner = (
            f"cd {req.machine.workdir} && {PROXY_EXPORTS} && "
            f"export HF_HOME=/storage/lkk/cache && "
            f"openclaw agent --local "
            f"--session-id {shlex.quote(req.session_key)} "
            f"--message {shlex.quote(task_prompt)} "
            f"--timeout {req.timeout}"
        )

        stop_event = asyncio.Event()
        session_task = asyncio.create_task(
            _tail_session_jsonl(queue, task_id, session_file, stop_event, container=container)
        )

        proc = await asyncio.create_subprocess_exec(
            "docker", "exec", container, "bash", "-c", inner,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.STDOUT,
        )

        async def _drain():
            async for raw in proc.stdout:
                line = raw.decode(errors="replace").rstrip()
                if line:
                    await emit("log", line)

        await asyncio.gather(_drain(), proc.wait())
        rc = proc.returncode

        await asyncio.sleep(3)
        stop_event.set()
        session_task.cancel()
        try:
            await session_task
        except asyncio.CancelledError:
            pass

        if rc == 0:
            await emit("done", f"Auto-Test finished: {req.model_id}")
        else:
            await emit("error", f"Auto-Test exited with code {rc}", level="error")

    except Exception as e:
        logger.exception("run_autotest_task failed")
        await queue.put(_evt(task_id, "error", str(e), level="error"))
    finally:
        await queue.put(None)


# ─────────────────────────── Auto-Quant task ──────────────────────────────

def _build_autoquant_prompt(req: "AutoQuantRequest") -> str:
    model_slug = req.model_id.replace("/", "_")
    output_dir = f"{req.output_dir}/{model_slug}-{req.scheme}"
    workspace_line = (
        f"- Shared workspace (auto_run output dir for this model, may contain model_info.json): {req.workspace_dir}\n"
        if req.workspace_dir else ""
    )

    return (
        f"You are an expert in LLM quantization using the Intel Auto-Round toolkit.\n"
        f"You MUST follow the skill instructions in: {req.skill_path}\n\n"
        f"Model: {req.model_id}\n"
        f"Quantization: {req.scheme} / {req.method}\n"
        f"Export format: {req.export_format}\n"
        f"Output directory: {output_dir}\n"
        f"Runtime device: {req.device}:{req.device_index}\n\n"
        f"System environment (concrete values for this machine):\n"
        f"- Shell: always use bash; use 'bash -c' explicitly in shell commands.\n"
        f"- Base Python: /kaokao/miniforge3/bin/python3  (has torch 2.9.0+xpu pre-installed)\n"
        f"- HuggingFace cache dir: HF_HOME=/storage/lkk/cache\n"
        f"- Network proxy: https_proxy=http://child-igk.intel.com:912  http_proxy=http://child-igk.intel.com:912\n"
        f"{workspace_line}"
    )


async def run_autoquant_task(task_id: str, req: "AutoQuantRequest"):
    queue = task_queues[task_id]

    async def emit(type_, msg, stage="autoquant", level="info", meta=None):
        await queue.put(_evt(task_id, type_, msg, stage, level, meta))

    try:
        await emit("status", f"Starting Auto-Quant: {req.model_id} [{req.scheme}/{req.method}]")
        container = req.machine.container

        session_file = f"{req.machine.session_dir}/{req.session_key}.jsonl"
        clear_proc = await asyncio.create_subprocess_exec(
            "docker", "exec", container,
            "bash", "-c", f"truncate -s 0 {shlex.quote(session_file)} 2>/dev/null || true",
            stdout=asyncio.subprocess.DEVNULL, stderr=asyncio.subprocess.DEVNULL,
        )
        await clear_proc.wait()

        task_prompt = _build_autoquant_prompt(req)
        inner = (
            f"cd {req.machine.workdir} && {PROXY_EXPORTS} && "
            f"export HF_HOME=/storage/lkk/cache && "
            f"openclaw agent --local "
            f"--session-id {shlex.quote(req.session_key)} "
            f"--message {shlex.quote(task_prompt)} "
            f"--timeout {req.timeout}"
        )

        stop_event = asyncio.Event()
        session_task = asyncio.create_task(
            _tail_session_jsonl(queue, task_id, session_file, stop_event, container=container)
        )

        proc = await asyncio.create_subprocess_exec(
            "docker", "exec", container, "bash", "-c", inner,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.STDOUT,
        )

        async def _drain():
            async for raw in proc.stdout:
                line = raw.decode(errors="replace").rstrip()
                if line:
                    await emit("log", line)

        await asyncio.gather(_drain(), proc.wait())
        rc = proc.returncode

        await asyncio.sleep(3)
        stop_event.set()
        session_task.cancel()
        try:
            await session_task
        except asyncio.CancelledError:
            pass

        if rc == 0:
            await emit("done", f"Auto-Quant finished: {req.model_id} [{req.scheme}]")
        else:
            await emit("error", f"Auto-Quant exited with code {rc}", level="error")

    except Exception as e:
        logger.exception("run_autoquant_task failed")
        await queue.put(_evt(task_id, "error", str(e), level="error"))
    finally:
        await queue.put(None)


# ─────────────────────────── Auto-Eval task ───────────────────────────────

def _build_autoeval_prompt(req: "AutoEvalRequest") -> str:
    return (
        f"You are an expert in LLM evaluation using vLLM and lm-evaluation-harness.\n"
        f"You MUST follow the skill instructions in: {req.skill_path}\n\n"
        f"Model: {req.model_path}\n"
        f"Evaluation tasks: {req.tasks}\n"
        f"Output directory: {req.output_dir}\n"
        f"Device: {req.device}:{req.device_index}\n"
        f"Batch size: {req.batch_size}\n"
        f"Max model length: {req.max_model_len}\n"
        f"GPU memory utilization: {req.gpu_memory_utilization}\n\n"
        f"System environment (concrete values for this machine):\n"
        f"- Shell: always use bash; use 'bash -c' explicitly in shell commands.\n"
        f"- Base Python: /kaokao/miniforge3/bin/python3  (has torch 2.9.0+xpu pre-installed)\n"
        f"- HuggingFace cache dir: HF_HOME=/storage/lkk/cache\n"
        f"- Network proxy: https_proxy=http://child-igk.intel.com:912  http_proxy=http://child-igk.intel.com:912\n"
    )


async def run_autoeval_task(task_id: str, req: "AutoEvalRequest"):
    queue = task_queues[task_id]

    async def emit(type_, msg, stage="autoeval", level="info", meta=None):
        await queue.put(_evt(task_id, type_, msg, stage, level, meta))

    try:
        await emit("status", f"Starting Auto-Eval: {req.model_path} tasks={req.tasks}")
        container = req.machine.container

        session_file = f"{req.machine.session_dir}/{req.session_key}.jsonl"
        clear_proc = await asyncio.create_subprocess_exec(
            "docker", "exec", container,
            "bash", "-c", f"truncate -s 0 {shlex.quote(session_file)} 2>/dev/null || true",
            stdout=asyncio.subprocess.DEVNULL, stderr=asyncio.subprocess.DEVNULL,
        )
        await clear_proc.wait()

        task_prompt = _build_autoeval_prompt(req)
        inner = (
            f"cd {req.machine.workdir} && {PROXY_EXPORTS} && "
            f"export HF_HOME=/storage/lkk/cache && "
            f"openclaw agent --local "
            f"--session-id {shlex.quote(req.session_key)} "
            f"--message {shlex.quote(task_prompt)} "
            f"--timeout {req.timeout}"
        )

        stop_event = asyncio.Event()
        session_task = asyncio.create_task(
            _tail_session_jsonl(queue, task_id, session_file, stop_event, container=container)
        )

        proc = await asyncio.create_subprocess_exec(
            "docker", "exec", container, "bash", "-c", inner,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.STDOUT,
        )

        async def _drain():
            async for raw in proc.stdout:
                line = raw.decode(errors="replace").rstrip()
                if line:
                    await emit("log", line)

        await asyncio.gather(_drain(), proc.wait())
        rc = proc.returncode

        await asyncio.sleep(3)
        stop_event.set()
        session_task.cancel()
        try:
            await session_task
        except asyncio.CancelledError:
            pass

        if rc == 0:
            await emit("done", f"Auto-Eval finished: {req.model_path}")
        else:
            await emit("error", f"Auto-Eval exited with code {rc}", level="error")

    except Exception as e:
        logger.exception("run_autoeval_task failed")
        await queue.put(_evt(task_id, "error", str(e), level="error"))
    finally:
        await queue.put(None)


# ─────────────────────────── Shared agent runner ──────────────────────────

async def _run_agent_stage(
    queue, task_id: str, session_key: str, prompt: str,
    workdir: str, timeout: int, stage_name: str,
    container: str = DOCKER_CONTAINER, session_dir: str = SESSION_DIR,
) -> bool:
    """Run one openclaw agent stage, streaming output into queue. Returns True on success."""
    session_file = f"{session_dir}/{session_key}.jsonl"
    clear_proc = await asyncio.create_subprocess_exec(
        "docker", "exec", container,
        "bash", "-c", f"truncate -s 0 {shlex.quote(session_file)} 2>/dev/null || true",
        stdout=asyncio.subprocess.DEVNULL, stderr=asyncio.subprocess.DEVNULL,
    )
    await clear_proc.wait()

    inner = (
        f"cd {workdir} && {PROXY_EXPORTS} && "
        f"export HF_HOME=/storage/lkk/cache && "
        f"openclaw agent --local "
        f"--session-id {shlex.quote(session_key)} "
        f"--message {shlex.quote(prompt)} "
        f"--timeout {timeout}"
    )

    stop_event = asyncio.Event()
    session_task = asyncio.create_task(
        _tail_session_jsonl(queue, task_id, session_file, stop_event, container=container)
    )

    proc = await asyncio.create_subprocess_exec(
        "docker", "exec", container, "bash", "-c", inner,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.STDOUT,
    )

    async def _drain():
        async for raw in proc.stdout:
            line = raw.decode(errors="replace").rstrip()
            if line:
                await queue.put(_evt(task_id, "log", line, stage=stage_name))

    await asyncio.gather(_drain(), proc.wait())
    rc = proc.returncode

    await asyncio.sleep(3)
    stop_event.set()
    session_task.cancel()
    try:
        await session_task
    except asyncio.CancelledError:
        pass

    return rc == 0


# ─────────────────────────── Pipeline task ────────────────────────────────

async def run_pipeline_task(task_id: str, req: PipelineRequest):
    queue = task_queues[task_id]
    model_slug = req.model_id.replace("/", "_")
    quant_output = f"{req.quant_output_dir}/{model_slug}-{req.scheme}"
    container   = req.machine.container
    workdir     = req.machine.workdir
    session_dir = req.machine.session_dir

    async def emit(type_, msg, stage="pipeline", level="info", meta=None):
        await queue.put(_evt(task_id, type_, msg, stage, level, meta))

    try:
        stages = []
        if req.run_autotest:  stages.append("Auto-Test")
        if req.run_autoquant: stages.append("Auto-Quant")
        if req.run_autoeval:  stages.append("Auto-Eval")
        if not stages:
            await emit("error", "No stages selected", level="error")
            return

        total = len(stages)
        step = 0
        await emit("status", f"Pipeline [{' → '.join(stages)}] | model={req.model_id}")

        # ── Stage: Auto-Test ──
        if req.run_autotest:
            step += 1
            await emit("status", f"[{step}/{total}] Auto-Test: {req.model_id}", stage="autotest")
            at_req = AutoTestRequest(
                model_id=req.model_id, device=req.device,
                device_index=req.device_index,
                output_dir=req.test_output_dir,
                skill_path=req.autorun_skill,
                session_key=f"{req.session_key}_test",
                timeout=req.timeout,
            )
            ok = await _run_agent_stage(
                queue, task_id,
                session_key=f"{req.session_key}_test",
                prompt=_build_autotest_prompt(at_req),
                workdir=workdir,
                timeout=req.timeout,
                stage_name="autotest",
                container=container, session_dir=session_dir,
            )
            if ok:
                await emit("done", f"Auto-Test done: {req.model_id}", stage="autotest")
            else:
                await emit("error", "Auto-Test failed — pipeline aborted", level="error")
                return

        # ── Stage: Auto-Quant ──
        if req.run_autoquant:
            step += 1
            await emit("status", f"[{step}/{total}] Auto-Quant: {req.model_id} [{req.scheme}/{req.method}]", stage="autoquant")
            aq_req = AutoQuantRequest(
                model_id=req.model_id, scheme=req.scheme, method=req.method,
                export_format=req.export_format, device=req.device,
                device_index=req.device_index, output_dir=req.quant_output_dir,
                workspace_dir=f"{req.test_output_dir}/{model_slug}",
                skill_path=req.autoquant_skill,
                session_key=f"{req.session_key}_quant",
                timeout=req.timeout,
            )
            ok = await _run_agent_stage(
                queue, task_id,
                session_key=f"{req.session_key}_quant",
                prompt=_build_autoquant_prompt(aq_req),
                workdir=workdir,
                timeout=req.timeout,
                stage_name="autoquant",
                container=container, session_dir=session_dir,
            )
            if ok:
                await emit("done", f"Auto-Quant done → {quant_output}", stage="autoquant")
            else:
                await emit("error", "Auto-Quant failed — pipeline aborted", level="error")
                return

        # ── Stage: Auto-Eval ──
        if req.run_autoeval:
            step += 1
            await emit("status", f"[{step}/{total}] Auto-Eval: {quant_output} tasks={req.tasks}", stage="autoeval")
            ae_req = AutoEvalRequest(
                model_path=quant_output,
                tasks=req.tasks,
                output_dir=f"{req.eval_output_dir}/{model_slug}",
                device=req.device,
                device_index=req.device_index,
                batch_size=req.batch_size,
                max_model_len=req.max_model_len,
                gpu_memory_utilization=req.gpu_memory_utilization,
                skill_path=req.skill_path,
                session_key=f"{req.session_key}_eval",
                timeout=req.timeout,
            )
            ok = await _run_agent_stage(
                queue, task_id,
                session_key=f"{req.session_key}_eval",
                prompt=_build_autoeval_prompt(ae_req),
                workdir=workdir,
                timeout=req.timeout,
                stage_name="autoeval",
                container=container, session_dir=session_dir,
            )
            if ok:
                await emit("done", f"Auto-Eval done: tasks={req.tasks}", stage="autoeval")
            else:
                await emit("error", "Auto-Eval failed", level="error")
                return

        await emit("done", f"Pipeline complete ✓  [{' → '.join(stages)}]")

    except Exception as e:
        logger.exception("run_pipeline_task failed")
        await queue.put(_evt(task_id, "error", str(e), level="error"))
    finally:
        await queue.put(None)


# ─────────────────────────── Excel-XLX task ──────────────────────────────

@app.post("/api/run-excel-xlx")
async def run_excel_xlx(req: ExcelXLXRequest):
    task_id = str(uuid.uuid4())
    task_queues[task_id] = asyncio.Queue()
    asyncio.create_task(run_excel_xlx_task(task_id, req))
    return {"task_id": task_id}


async def run_excel_xlx_task(task_id: str, req: ExcelXLXRequest):
    queue = task_queues[task_id]

    async def emit(type_, msg, stage="excel_xlx", level="info", meta=None):
        await queue.put(_evt(task_id, type_, msg, stage, level, meta))

    try:
        container = req.machine.container

        # Build the message: combine instructions + optional file path + optional skill
        message = req.instructions.strip()
        if req.file_path.strip():
            message = f"{message}\n\nExcel file path: {req.file_path.strip()}"
        if req.skill_path.strip():
            message = f"You must follow the skill instructions in: {req.skill_path}\n\n{message}"

        await emit("status", f"Starting Excel-XLX agent: {message[:80]!r}")

        inner = (
            f"cd {shlex.quote(req.machine.workdir)} && "
            f"openclaw agent --local --agent main "
            f"--message {shlex.quote(message)} "
            f"--timeout {req.timeout}"
        )

        # Pass MINIMAX_API_KEY and proxy as -e args to docker exec (mirrors the manual command)
        docker_args = ["docker", "exec"]
        if req.machine.minimax_key:
            docker_args += ["-e", f"MINIMAX_API_KEY={req.machine.minimax_key}"]
        docker_args += [
            "-e", "https_proxy=http://child-igk.intel.com:912",
            "-e", "http_proxy=http://child-igk.intel.com:912",
            container,
            "bash", "-lc", inner,
        ]

        # openclaw writes to UUID-named files, not --session-id named files.
        # Use _watch_and_tail_session to detect whichever file actually gets written.
        stop_event = asyncio.Event()
        session_task = asyncio.create_task(
            _watch_and_tail_session(queue, task_id, req.machine.session_dir, stop_event, container=container)
        )

        proc = await asyncio.create_subprocess_exec(
            *docker_args,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.STDOUT,
        )

        async def _drain():
            async for raw in proc.stdout:
                line = raw.decode(errors="replace").rstrip()
                if line:
                    await emit("log", line)

        await asyncio.gather(_drain(), proc.wait())
        rc = proc.returncode

        await asyncio.sleep(3)
        stop_event.set()
        session_task.cancel()
        try:
            await session_task
        except asyncio.CancelledError:
            pass

        if rc == 0:
            await emit("done", "Excel-XLX agent finished successfully")
        else:
            await emit("error", f"Excel-XLX agent exited with code {rc}", level="error")

    except Exception as e:
        logger.exception("run_excel_xlx_task failed")
        await queue.put(_evt(task_id, "error", str(e), level="error"))
    finally:
        await queue.put(None)


@app.post("/api/replay-session")
async def replay_session(
    session_dir: str = SESSION_DIR,
    session_file: str = "",
    container: str = DOCKER_CONTAINER,
):
    """
    Replay transcript events from an existing session JSONL file.
    If session_file is empty, uses the most recently modified non-empty .jsonl in session_dir.
    Returns a task_id; stream events via /api/stream/{task_id}.
    """
    task_id = str(uuid.uuid4())
    task_queues[task_id] = asyncio.Queue()

    async def _replay():
        queue = task_queues[task_id]
        try:
            if not session_file:
                # Find the most recently modified non-empty .jsonl
                find_cmd = (
                    f"find {shlex.quote(session_dir)} -maxdepth 1 -name '*.jsonl' "
                    f"-exec stat -c '%Y %n' {{}} + 2>/dev/null | sort -rn | head -5"
                )
                p = await asyncio.create_subprocess_exec(
                    "docker", "exec", container, "bash", "-c", find_cmd,
                    stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.DEVNULL,
                )
                out, _ = await p.communicate()
                target = None
                for ln in out.decode().splitlines():
                    parts = ln.split(None, 1)
                    if len(parts) == 2:
                        candidate = parts[1].strip()
                        # Skip empty files
                        sz_p = await asyncio.create_subprocess_exec(
                            "docker", "exec", container, "bash", "-c",
                            f"stat -c%s {shlex.quote(candidate)} 2>/dev/null",
                            stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.DEVNULL,
                        )
                        sz_out, _ = await sz_p.communicate()
                        if sz_out.strip() and int(sz_out.strip()) > 0:
                            target = candidate
                            break
                if not target:
                    await queue.put(_evt(task_id, "error", "No non-empty session file found", level="error"))
                    return
            else:
                target = session_file

            await queue.put(_evt(task_id, "status", f"Replaying session: {target}"))

            # Read entire file and emit transcript events
            cat_proc = await asyncio.create_subprocess_exec(
                "docker", "exec", container, "cat", target,
                stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.DEVNULL,
            )
            out, _ = await cat_proc.communicate()
            for line in out.decode(errors="replace").splitlines():
                line = line.strip()
                if not line:
                    continue
                for sub_type, text in _parse_session_line(line):
                    await queue.put(_evt(task_id, "transcript", text,
                                        meta={"sub_type": sub_type}))
            await queue.put(_evt(task_id, "done", f"Replay complete: {target}"))
        except Exception as e:
            await queue.put(_evt(task_id, "error", str(e), level="error"))
        finally:
            await queue.put(None)

    asyncio.create_task(_replay())
    return {"task_id": task_id}


@app.get("/api/list-sessions")
async def list_sessions(
    session_dir: str = SESSION_DIR,
    container: str = DOCKER_CONTAINER,
):
    """List all session JSONL files in session_dir with their sizes and modification times."""
    cmd = (
        f"find {shlex.quote(session_dir)} -maxdepth 1 -name '*.jsonl' "
        f"-exec stat -c '%Y %s %n' {{}} + 2>/dev/null | sort -rn"
    )
    p = await asyncio.create_subprocess_exec(
        "docker", "exec", container, "bash", "-c", cmd,
        stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.DEVNULL,
    )
    out, _ = await p.communicate()
    sessions = []
    for ln in out.decode().splitlines():
        parts = ln.split(None, 2)
        if len(parts) == 3:
            mtime, size, path = int(parts[0]), int(parts[1]), parts[2].strip()
            sessions.append({
                "path": path,
                "name": path.rsplit("/", 1)[-1],
                "size": size,
                "mtime": mtime,
            })
    return {"sessions": sessions}


@app.get("/api/read-xlsx")
async def read_xlsx(path: str, container: str = DOCKER_CONTAINER):
    """Read an xlsx file from inside Docker and return it as markdown tables."""
    if not path.startswith("/"):
        raise HTTPException(status_code=400, detail="path must be absolute")
    script = (
        "import openpyxl, json, sys; "
        f"wb = openpyxl.load_workbook({repr(path)}, read_only=True, data_only=True); "
        "result = {}; "
        "[result.update({name: [list(r) for r in wb[name].iter_rows(values_only=True)]}) for name in wb.sheetnames]; "
        "print(json.dumps(result, default=str))"
    )
    proc = await asyncio.create_subprocess_exec(
        "docker", "exec", container, "python3", "-c", script,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE,
    )
    stdout, stderr = await proc.communicate()
    if proc.returncode != 0:
        raise HTTPException(status_code=500, detail=stderr.decode(errors="replace")[:500])
    data = json.loads(stdout.decode())
    md = ""
    for sheet, rows in data.items():
        md += f"## {sheet}\n\n"
        if rows:
            header = [str(c) if c is not None else "" for c in rows[0]]
            md += "| " + " | ".join(header) + " |\n"
            md += "| " + " | ".join("---" for _ in header) + " |\n"
            for row in rows[1:]:
                cells = [str(c) if c is not None else "" for c in row]
                md += "| " + " | ".join(cells) + " |\n"
        md += "\n"
    return {"markdown": md}


@app.get("/api/image")
async def get_image(path: str, container: str = DOCKER_CONTAINER):
    """Serve an image file from inside the Docker container."""
    if not path.startswith("/"):
        raise HTTPException(status_code=400, detail="path must be absolute")
    proc = await asyncio.create_subprocess_exec(
        "docker", "exec", container, "cat", path,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE,
    )
    data, _ = await proc.communicate()
    if proc.returncode != 0 or not data:
        raise HTTPException(status_code=404, detail="Image not found in container")
    ext = path.rsplit(".", 1)[-1].lower() if "." in path else ""
    ct = {"png": "image/png", "jpg": "image/jpeg", "jpeg": "image/jpeg",
          "gif": "image/gif", "webp": "image/webp"}.get(ext, "application/octet-stream")
    return Response(content=data, media_type=ct)


@app.get("/api/stream/{task_id}")
async def stream(task_id: str):
    queue = task_queues.get(task_id)
    if not queue:
        raise HTTPException(status_code=404, detail="task not found")

    async def generator():
        try:
            while True:
                try:
                    item = await asyncio.wait_for(queue.get(), timeout=15.0)
                except asyncio.TimeoutError:
                    yield ": heartbeat\n\n"
                    continue
                if item is None:
                    yield "data: [DONE]\n\n"
                    break
                yield f"data: {json.dumps(item, ensure_ascii=False)}\n\n"
        finally:
            task_queues.pop(task_id, None)

    return StreamingResponse(generator(), media_type="text/event-stream")
